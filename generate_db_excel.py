# -*- coding: utf-8 -*-
"""
数据库表结构导出为Excel
需要安装: pip install openpyxl
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 创建工作簿
wb = Workbook()

# 定义样式
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 表结构定义
tables = [
    {
        'name': 'user-用户表',
        'columns': [
            ('字段名', '数据类型', '是否为空', '是否主键', '默认值', '描述'),
            ('id', 'BIGINT', '否', '是', 'NULL', '用户ID，主键自增'),
            ('openid', 'VARCHAR(64)', '是', '否', 'NULL', '微信openid，唯一'),
            ('nickname', 'VARCHAR(64)', '否', '否', 'NULL', '用户昵称'),
            ('avatar_url', 'VARCHAR(255)', '是', '否', 'NULL', '头像URL'),
            ('phone', 'VARCHAR(32)', '是', '否', 'NULL', '手机号'),
            ('status', 'TINYINT', '否', '否', '1', '状态（1正常 0禁用）'),
            ('password', 'VARCHAR(64)', '是', '否', 'NULL', '密码（MD5加密）'),
            ('created_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '创建时间'),
            ('updated_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '更新时间'),
            ('is_deleted', 'TINYINT', '否', '否', '0', '是否删除（0否 1是）'),
        ]
    },
    {
        'name': 'admin_user-管理员表',
        'columns': [
            ('字段名', '数据类型', '是否为空', '是否主键', '默认值', '描述'),
            ('id', 'BIGINT', '否', '是', 'NULL', '管理员ID，主键自增'),
            ('username', 'VARCHAR(64)', '否', '否', 'NULL', '用户名，唯一'),
            ('password', 'VARCHAR(64)', '否', '否', 'NULL', '密码'),
            ('real_name', 'VARCHAR(64)', '否', '否', 'NULL', '真实姓名'),
            ('role_name', 'VARCHAR(32)', '否', '否', 'AUDITOR', '角色名称'),
            ('status', 'TINYINT', '否', '否', '1', '状态（1正常 0禁用）'),
            ('created_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '创建时间'),
            ('updated_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '更新时间'),
            ('is_deleted', 'TINYINT', '否', '否', '0', '是否删除'),
        ]
    },
    {
        'name': 'category-商品分类表',
        'columns': [
            ('字段名', '数据类型', '是否为空', '是否主键', '默认值', '描述'),
            ('id', 'BIGINT', '否', '是', 'NULL', '分类ID，主键自增'),
            ('name', 'VARCHAR(64)', '否', '否', 'NULL', '分类名称'),
            ('sort', 'INT', '否', '否', '0', '排序值'),
            ('status', 'TINYINT', '否', '否', '1', '状态（1正常 0禁用）'),
            ('created_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '创建时间'),
            ('updated_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '更新时间'),
            ('is_deleted', 'TINYINT', '否', '否', '0', '是否删除'),
        ]
    },
    {
        'name': 'goods-商品表',
        'columns': [
            ('字段名', '数据类型', '是否为空', '是否主键', '默认值', '描述'),
            ('id', 'BIGINT', '否', '是', 'NULL', '商品ID，主键自增'),
            ('user_id', 'BIGINT', '否', '否', 'NULL', '发布用户ID，外键'),
            ('category_id', 'BIGINT', '否', '否', 'NULL', '分类ID，外键'),
            ('title', 'VARCHAR(128)', '否', '否', 'NULL', '商品标题'),
            ('description', 'TEXT', '否', '否', 'NULL', '商品描述'),
            ('price', 'DECIMAL(10,2)', '否', '否', 'NULL', '商品价格'),
            ('cover_image', 'VARCHAR(255)', '否', '否', 'NULL', '封面图片URL'),
            ('status', 'VARCHAR(32)', '否', '否', 'PENDING', '商品状态'),
            ('reject_reason', 'VARCHAR(255)', '是', '否', 'NULL', '驳回原因'),
            ('view_count', 'INT', '否', '否', '0', '浏览次数'),
            ('favorite_count', 'INT', '否', '否', '0', '收藏次数'),
            ('comment_count', 'INT', '否', '否', '0', '留言次数'),
            ('created_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '创建时间'),
            ('updated_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '更新时间'),
            ('is_deleted', 'TINYINT', '否', '否', '0', '是否删除'),
        ]
    },
    {
        'name': 'goods_image-商品图片表',
        'columns': [
            ('字段名', '数据类型', '是否为空', '是否主键', '默认值', '描述'),
            ('id', 'BIGINT', '否', '是', 'NULL', '图片ID，主键自增'),
            ('goods_id', 'BIGINT', '否', '否', 'NULL', '商品ID，外键'),
            ('image_url', 'VARCHAR(255)', '否', '否', 'NULL', '图片URL'),
            ('sort', 'INT', '否', '否', '1', '排序值'),
            ('created_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '创建时间'),
            ('updated_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '更新时间'),
            ('is_deleted', 'TINYINT', '否', '否', '0', '是否删除'),
        ]
    },
    {
        'name': 'goods_favorite-商品收藏表',
        'columns': [
            ('字段名', '数据类型', '是否为空', '是否主键', '默认值', '描述'),
            ('id', 'BIGINT', '否', '是', 'NULL', '收藏ID，主键自增'),
            ('user_id', 'BIGINT', '否', '否', 'NULL', '用户ID，外键'),
            ('goods_id', 'BIGINT', '否', '否', 'NULL', '商品ID，外键'),
            ('created_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '创建时间'),
            ('updated_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '更新时间'),
            ('is_deleted', 'TINYINT', '否', '否', '0', '是否删除'),
        ]
    },
    {
        'name': 'goods_comment-商品留言表',
        'columns': [
            ('字段名', '数据类型', '是否为空', '是否主键', '默认值', '描述'),
            ('id', 'BIGINT', '否', '是', 'NULL', '留言ID，主键自增'),
            ('goods_id', 'BIGINT', '否', '否', 'NULL', '商品ID，外键'),
            ('user_id', 'BIGINT', '否', '否', 'NULL', '用户ID，外键'),
            ('content', 'VARCHAR(500)', '否', '否', 'NULL', '留言内容'),
            ('created_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '创建时间'),
            ('updated_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '更新时间'),
            ('is_deleted', 'TINYINT', '否', '否', '0', '是否删除'),
        ]
    },
    {
        'name': 'trade_order-交易订单表',
        'columns': [
            ('字段名', '数据类型', '是否为空', '是否主键', '默认值', '描述'),
            ('id', 'BIGINT', '否', '是', 'NULL', '订单ID，主键自增'),
            ('order_no', 'VARCHAR(64)', '否', '否', 'NULL', '订单编号，唯一'),
            ('goods_id', 'BIGINT', '否', '否', 'NULL', '商品ID，外键'),
            ('buyer_id', 'BIGINT', '否', '否', 'NULL', '买家用户ID，外键'),
            ('seller_id', 'BIGINT', '否', '否', 'NULL', '卖家用户ID，外键'),
            ('amount', 'DECIMAL(10,2)', '否', '否', 'NULL', '订单金额'),
            ('note', 'VARCHAR(500)', '是', '否', 'NULL', '订单备注'),
            ('order_status', 'VARCHAR(32)', '否', '否', 'PENDING_PAYMENT', '订单状态'),
            ('pay_status', 'VARCHAR(32)', '否', '否', 'UNPAID', '支付状态'),
            ('buyer_confirmed', 'TINYINT', '否', '否', '0', '买家已确认'),
            ('seller_confirmed', 'TINYINT', '否', '否', '0', '卖家已确认'),
            ('paid_at', 'DATETIME', '是', '否', 'NULL', '支付时间'),
            ('finished_at', 'DATETIME', '是', '否', 'NULL', '完成时间'),
            ('created_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '创建时间'),
            ('updated_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '更新时间'),
            ('is_deleted', 'TINYINT', '否', '否', '0', '是否删除'),
        ]
    },
    {
        'name': 'goods_report-商品举报表',
        'columns': [
            ('字段名', '数据类型', '是否为空', '是否主键', '默认值', '描述'),
            ('id', 'BIGINT', '否', '是', 'NULL', '举报ID，主键自增'),
            ('goods_id', 'BIGINT', '否', '否', 'NULL', '商品ID，外键'),
            ('reporter_id', 'BIGINT', '否', '否', 'NULL', '举报人ID，外键'),
            ('reason', 'VARCHAR(255)', '否', '否', 'NULL', '举报原因'),
            ('content', 'VARCHAR(500)', '是', '否', 'NULL', '举报详细说明'),
            ('status', 'VARCHAR(32)', '否', '否', 'PENDING', '处理状态'),
            ('handler_id', 'BIGINT', '是', '否', 'NULL', '处理人ID，外键'),
            ('handle_result', 'VARCHAR(500)', '是', '否', 'NULL', '处理结果'),
            ('handled_at', 'DATETIME', '是', '否', 'NULL', '处理时间'),
            ('created_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '创建时间'),
            ('updated_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '更新时间'),
            ('is_deleted', 'TINYINT', '否', '否', '0', '是否删除'),
        ]
    },
    {
        'name': 'notice-平台公告表',
        'columns': [
            ('字段名', '数据类型', '是否为空', '是否主键', '默认值', '描述'),
            ('id', 'BIGINT', '否', '是', 'NULL', '公告ID，主键自增'),
            ('title', 'VARCHAR(128)', '否', '否', 'NULL', '公告标题'),
            ('content', 'TEXT', '否', '否', 'NULL', '公告内容'),
            ('cover_url', 'VARCHAR(255)', '是', '否', 'NULL', '封面图片URL'),
            ('status', 'TINYINT', '否', '否', '1', '状态（1上线 0下线）'),
            ('publish_admin_id', 'BIGINT', '是', '否', 'NULL', '发布管理员ID'),
            ('published_at', 'DATETIME', '是', '否', 'NULL', '发布时间'),
            ('created_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '创建时间'),
            ('updated_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '更新时间'),
            ('is_deleted', 'TINYINT', '否', '否', '0', '是否删除'),
        ]
    },
    {
        'name': 'goods_audit-商品审核记录表',
        'columns': [
            ('字段名', '数据类型', '是否为空', '是否主键', '默认值', '描述'),
            ('id', 'BIGINT', '否', '是', 'NULL', '审核记录ID，主键自增'),
            ('goods_id', 'BIGINT', '否', '否', 'NULL', '商品ID，外键'),
            ('admin_id', 'BIGINT', '否', '否', 'NULL', '审核管理员ID，外键'),
            ('result', 'VARCHAR(32)', '否', '否', 'NULL', '审核结果'),
            ('reason', 'VARCHAR(255)', '是', '否', 'NULL', '驳回原因'),
            ('created_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '审核时间'),
            ('updated_at', 'DATETIME', '否', '否', 'CURRENT_TIMESTAMP', '更新时间'),
            ('is_deleted', 'TINYINT', '否', '否', '0', '是否删除'),
        ]
    },
]

# 删除默认工作表
wb.remove(wb.active)

# 为每个表创建一个工作表
for table in tables:
    ws = wb.create_sheet(title=table['name'])
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 35

    for row_idx, row_data in enumerate(table['columns'], 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill

# 创建表清单工作表
ws_index = wb.create_sheet(title='表清单', index=0)
ws_index.column_dimensions['A'].width = 25
ws_index.column_dimensions['B'].width = 50

ws_index.cell(row=1, column=1, value='序号').font = header_font
ws_index.cell(row=1, column=1, value='序号').fill = header_fill
ws_index.cell(row=1, column=1).alignment = header_alignment
ws_index.cell(row=1, column=1).border = thin_border

ws_index.cell(row=1, column=2, value='表名').font = header_font
ws_index.cell(row=1, column=2, value='表名').fill = header_fill
ws_index.cell(row=1, column=2).alignment = header_alignment
ws_index.cell(row=1, column=2).border = thin_border

for idx, table in enumerate(tables, 1):
    ws_index.cell(row=idx+1, column=1, value=idx).alignment = cell_alignment
    ws_index.cell(row=idx+1, column=1).border = thin_border
    ws_index.cell(row=idx+1, column=2, value=table['name']).alignment = cell_alignment
    ws_index.cell(row=idx+1, column=2).border = thin_border

# 保存文件
output_path = os.path.join(os.path.dirname(__file__), '数据库表结构.xlsx')
wb.save(output_path)
print(f'Excel文件已生成: {output_path}')