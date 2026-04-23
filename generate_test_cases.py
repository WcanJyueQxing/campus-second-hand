# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_sheet(wb, sheet_name, test_cases, title_color="4472C4"):
    ws = wb.create_sheet(title=sheet_name)
    headers = ["用例编号", "用例名称", "功能模块", "优先级", "前置条件", "测试步骤", "预期结果"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=title_color, end_color=title_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 35
    return ws

# ========== 小程序用户端测试用例 ==========

# 1. 用户认证模块
user_auth_cases = [
    ["UC001", "微信用户登录-新用户", "用户认证", "高", "微信小程序已打开", "1. 点击微信登录\n2. 授权获取用户信息", "成功登录并创建账号，跳转首页"],
    ["UC002", "微信用户登录-老用户", "用户认证", "高", "用户已通过微信登录过", "1. 点击微信登录\n2. 授权获取用户信息", "成功登录并跳转首页，显示用户昵称"],
    ["UC003", "账号密码登录-正确账号", "用户认证", "高", "用户已注册且账号正常", "1. 切换到账号登录\n2. 输入正确账号和密码\n3. 点击登录", "登录成功，跳转首页"],
    ["UC004", "账号密码登录-错误密码", "用户认证", "中", "用户已注册", "1. 切换到账号登录\n2. 输入正确账号和错误密码\n3. 点击登录", "提示\"密码错误\"，不允许登录"],
    ["UC005", "账号密码登录-未注册账号", "用户认证", "中", "账号未注册", "1. 切换到账号登录\n2. 输入未注册的账号和任意密码\n3. 点击登录", "提示\"账号不存在，请先注册\""],
    ["UC006", "记住密码功能", "用户认证", "低", "用户已登录并勾选记住密码", "1. 登录时勾选记住密码\n2. 退出登录\n3. 重新打开登录页", "自动填充上次登录的账号密码"],
    ["UC007", "退出登录", "用户认证", "中", "用户已登录", "1. 点击个人中心\n2. 点击退出登录", "清除登录状态，返回登录页"],
]

# 2. 商品浏览模块
goods_browse_cases = [
    ["GD001", "首页商品列表展示", "商品浏览", "高", "系统有已审核商品", "1. 打开小程序\n2. 首页默认加载", "显示商品列表，包含商品图片、名称、价格"],
    ["GD002", "商品分类筛选", "商品浏览", "高", "存在多个分类的商品", "1. 点击分类标签\n2. 选择某一分类", "显示该分类下的所有商品"],
    ["GD003", "商品详情查看", "商品浏览", "高", "存在已审核商品", "1. 点击商品卡片\n2. 进入商品详情页", "显示商品图片、价格、描述、卖家信息"],
    ["GD004", "商品搜索-关键词匹配", "商品浏览", "高", "存在匹配商品", "1. 点击搜索框\n2. 输入商品名称关键词\n3. 点击搜索", "显示包含关键词的商品列表"],
    ["GD005", "商品搜索-无匹配结果", "商品浏览", "中", "关键词无匹配商品", "1. 输入不存在的商品名称\n2. 点击搜索", "提示\"未找到相关商品\""],
]

# 3. 商品发布模块
goods_publish_cases = [
    ["PD001", "发布商品-必填项验证", "商品发布", "高", "用户已登录", "1. 点击发布按钮\n2. 不填写任何信息直接提交", "提示必填项不能为空"],
    ["PD002", "发布商品-成功提交", "商品发布", "高", "用户已登录", "1. 填写商品标题、描述、价格\n2. 选择分类\n3. 上传商品图片\n4. 点击提交", "发布成功，商品进入待审核状态"],
    ["PD003", "发布商品-图片上传", "商品发布", "中", "用户已登录", "1. 选择商品图片\n2. 上传1-9张图片", "图片上传成功并显示预览"],
    ["PD004", "我的商品列表", "商品发布", "中", "用户已发布过商品", "1. 进入个人中心\n2. 点击我的发布", "显示用户发布的所有商品及状态"],
    ["PD005", "编辑商品信息", "商品发布", "中", "用户有未售出商品", "1. 进入我的发布\n2. 点击编辑\n3. 修改商品信息\n4. 保存", "商品信息更新成功"],
    ["PD006", "下架商品", "商品发布", "中", "用户有在售商品", "1. 进入我的发布\n2. 点击下架", "商品状态变为已下架，从列表移除"],
]

# 4. 商品购买模块
goods_purchase_cases = [
    ["PC001", "创建订单-正常流程", "商品购买", "高", "存在可购买的商品", "1. 查看商品详情\n2. 点击立即下单\n3. 确认订单信息\n4. 提交订单", "订单创建成功，跳转支付页面"],
    ["PC002", "创建订单-购买自己商品", "商品购买", "高", "用户查看自己发布的商品", "1. 查看自己发布的商品详情\n2. 点击立即下单", "隐藏下单按钮或提示\"不能购买自己发布的商品\""],
    ["PC003", "订单支付", "商品购买", "高", "有待支付订单", "1. 进入我的订单\n2. 点击待支付订单\n3. 点击支付", "支付成功，订单状态变为待发货"],
    ["PC004", "取消订单-待支付状态", "商品购买", "中", "有待支付订单", "1. 进入我的订单\n2. 点击取消订单", "订单取消成功，商品恢复可购买状态"],
    ["PC005", "确认收货", "商品购买", "高", "有已发货订单", "1. 收到货物后\n2. 进入我的订单\n3. 点击确认收货", "订单完成，金额打给卖家"],
]

# 5. 收藏功能模块
favorite_cases = [
    ["FC001", "添加商品收藏", "收藏功能", "中", "存在可购买商品", "1. 进入商品详情\n2. 点击收藏按钮", "收藏成功，按钮变为已收藏状态"],
    ["FC002", "取消商品收藏", "收藏功能", "中", "用户有收藏商品", "1. 进入商品详情\n2. 点击取消收藏", "取消收藏成功，按钮恢复未收藏状态"],
    ["FC003", "查看收藏列表", "收藏功能", "中", "用户有收藏记录", "1. 进入个人中心\n2. 点击我的收藏", "显示所有收藏的商品列表"],
]

# 6. 评价功能模块
comment_cases = [
    ["CM001", "商品评论-成功提交", "评价功能", "中", "用户已购买商品", "1. 进入已完成的订单\n2. 点击去评价\n3. 输入评论内容\n4. 提交", "评论成功，显示在商品详情页"],
    ["CM002", "查看商品评论", "评价功能", "中", "商品有评论", "1. 进入商品详情页\n2. 滑动到评论区", "显示该商品的所有评论"],
]

# 7. 举报功能模块
report_cases = [
    ["RP001", "举报商品-正常提交", "举报功能", "中", "存在可举报商品", "1. 进入商品详情\n2. 点击举报\n3. 选择举报原因\n4. 提交", "举报成功，等待后台审核"],
    ["RP002", "举报商品-重复举报", "举报功能", "低", "用户已举报过该商品", "1. 对已举报商品再次举报", "提示\"您已举报过该商品\""],
]

# 8. 订单管理模块
order_cases = [
    ["OM001", "查看我的订单列表", "订单管理", "高", "用户有订单", "1. 进入个人中心\n2. 点击我的订单", "显示所有订单，按状态分类"],
    ["OM002", "订单状态筛选", "订单管理", "中", "用户有多种状态订单", "1. 在订单列表点击某一状态标签", "只显示该状态下的订单"],
    ["OM003", "订单详情查看", "订单管理", "中", "存在订单", "1. 点击某一订单", "显示订单详细信息，包含商品、金额、状态等"],
]

# 9. 个人信息模块
profile_cases = [
    ["PI001", "查看个人信息", "个人信息", "中", "用户已登录", "1. 进入个人中心", "显示用户昵称、头像、手机号等信息"],
    ["PI002", "编辑个人信息", "个人信息", "中", "用户已登录", "1. 点击编辑资料\n2. 修改昵称或头像\n3. 保存", "个人信息更新成功"],
]

# ========== 管理后台测试用例 ==========

# 10. 管理员登录模块
admin_auth_cases = [
    ["AM001", "管理员登录-正确账号", "管理员登录", "高", "管理员账号存在", "1. 输入正确的用户名和密码\n2. 点击登录", "登录成功，进入管理后台首页"],
    ["AM002", "管理员登录-错误密码", "管理员登录", "中", "管理员账号存在", "1. 输入正确的用户名\n2. 输入错误的密码\n3. 点击登录", "提示\"账号或密码错误\""],
    ["AM003", "管理员登录-禁用账号", "管理员登录", "中", "账号已被禁用", "1. 使用已禁用的账号登录", "提示\"账号不存在或已禁用\""],
]

# 11. 数据统计模块
dashboard_cases = [
    ["DM001", "查看数据统计", "数据统计", "高", "管理员已登录", "1. 登录管理后台\n2. 进入数据统计页面", "显示用户总数、商品总数、订单总数等"],
    ["DM002", "查看近7日趋势", "数据统计", "中", "存在近7日数据", "1. 进入数据统计页面", "显示近7日新增用户、商品、订单的趋势图"],
    ["DM003", "查看商品分类分布", "数据统计", "中", "存在商品数据", "1. 进入数据统计页面", "显示各分类商品的饼图分布"],
]

# 12. 商品审核模块
goods_audit_cases = [
    ["GA001", "查看待审核商品列表", "商品审核", "高", "有待审核商品", "1. 进入商品审核页面", "显示所有待审核商品列表"],
    ["GA002", "审核商品-通过", "商品审核", "高", "有待审核商品", "1. 选择待审核商品\n2. 点击通过", "商品状态变为已通过，可在前台展示"],
    ["GA003", "审核商品-拒绝", "商品审核", "高", "有待审核商品", "1. 选择待审核商品\n2. 点击拒绝\n3. 输入拒绝原因", "商品被拒绝，卖家收到拒绝通知"],
    ["GA004", "下架商品", "商品审核", "中", "存在已上架商品", "1. 找到已上架商品\n2. 点击下架", "商品下架，不再前台展示"],
]

# 13. 用户管理模块
user_manage_cases = [
    ["UM001", "查看用户列表", "用户管理", "高", "存在注册用户", "1. 进入用户管理页面", "显示所有用户列表，包含昵称、手机号、状态"],
    ["UM002", "禁用用户账号", "用户管理", "中", "存在正常用户", "1. 选择某一用户\n2. 点击禁用", "用户被禁用，无法登录"],
    ["UM003", "启用用户账号", "用户管理", "中", "存在已禁用用户", "1. 选择某一已禁用用户\n2. 点击启用", "用户恢复使用，可以正常登录"],
]

# 14. 分类管理模块
category_manage_cases = [
    ["CGM001", "查看分类列表", "分类管理", "中", "存在商品分类", "1. 进入分类管理页面", "显示所有商品分类列表"],
    ["CGM002", "添加商品分类", "分类管理", "中", "管理员已登录", "1. 点击添加分类\n2. 输入分类名称和排序\n3. 保存", "新分类创建成功"],
    ["CGM003", "编辑商品分类", "分类管理", "中", "存在可编辑分类", "1. 点击编辑分类\n2. 修改分类信息\n3. 保存", "分类信息更新成功"],
    ["CGM004", "删除商品分类", "分类管理", "中", "存在空分类", "1. 选择无商品的分类\n2. 点击删除", "分类删除成功"],
    ["CGM005", "删除有商品的分类", "分类管理", "低", "分类下有商品", "1. 选择有商品的分类\n2. 点击删除", "提示\"该分类下有商品，无法删除\""],
]

# 15. 订单监管模块
order_manage_cases = [
    ["OG001", "查看所有订单", "订单监管", "高", "存在订单数据", "1. 进入订单监管页面", "显示所有用户订单列表"],
    ["OG002", "订单状态筛选", "订单监管", "中", "存在多种状态订单", "1. 选择某一订单状态", "只显示该状态下的所有订单"],
    ["OG003", "查看订单详情", "订单监管", "中", "存在订单", "1. 点击某一订单", "显示订单详细信息，包含买家、卖家、商品、金额等"],
]

# 16. 举报处理模块
report_manage_cases = [
    ["RR001", "查看待处理举报", "举报处理", "高", "存在待处理举报", "1. 进入举报处理页面", "显示所有待处理的举报列表"],
    ["RR002", "处理举报-通过", "举报处理", "高", "存在待处理举报", "1. 选择举报信息\n2. 核实后点击通过\n3. 输入处理结果", "举报成立，商品被下架或删除"],
    ["RR003", "处理举报-驳回", "举报处理", "中", "存在待处理举报", "1. 选择举报信息\n2. 核实后点击驳回\n3. 输入驳回原因", "举报被驳回，商品恢复展示"],
]

# 17. 公告管理模块
notice_manage_cases = [
    ["NM001", "查看公告列表", "公告管理", "中", "存在公告", "1. 进入公告管理页面", "显示所有公告列表"],
    ["NM002", "发布新公告", "公告管理", "中", "管理员已登录", "1. 点击发布公告\n2. 输入标题和内容\n3. 选择发布范围\n4. 保存", "公告发布成功，用户可见"],
    ["NM003", "编辑公告", "公告管理", "中", "存在可编辑公告", "1. 选择某一公告\n2. 点击编辑\n3. 修改内容\n4. 保存", "公告更新成功"],
    ["NM004", "删除公告", "公告管理", "中", "存在公告", "1. 选择某一公告\n2. 点击删除", "公告删除成功，用户不可见"],
]

# 18. 文件上传模块
file_upload_cases = [
    ["FU001", "图片上传-正常", "文件上传", "中", "管理员已登录", "1. 进入文件管理\n2. 选择图片上传", "图片上传成功，显示在列表中"],
    ["FU002", "图片上传-格式错误", "文件上传", "低", "管理员已登录", "1. 选择非图片文件上传", "提示\"只支持jpg、png等图片格式\""],
]

# 删除默认工作表
del wb['Sheet']

# 创建分类工作表 - 小程序用户端
create_sheet(wb, "1-用户认证", user_auth_cases, "4472C4")
create_sheet(wb, "2-商品浏览", goods_browse_cases, "548235")
create_sheet(wb, "3-商品发布", goods_publish_cases, "C65911")
create_sheet(wb, "4-商品购买", goods_purchase_cases, "7030A0")
create_sheet(wb, "5-收藏功能", favorite_cases, "00B0F0")
create_sheet(wb, "6-评价功能", comment_cases, "FF0000")
create_sheet(wb, "7-举报功能", report_cases, "BF8F00")
create_sheet(wb, "8-订单管理", order_cases, "92D050")
create_sheet(wb, "9-个人信息", profile_cases, "FFC000")

# 创建分类工作表 - 管理后台
create_sheet(wb, "10-管理员登录", admin_auth_cases, "4472C4")
create_sheet(wb, "11-数据统计", dashboard_cases, "002060")
create_sheet(wb, "12-商品审核", goods_audit_cases, "548235")
create_sheet(wb, "13-用户管理", user_manage_cases, "C65911")
create_sheet(wb, "14-分类管理", category_manage_cases, "7030A0")
create_sheet(wb, "15-订单监管", order_manage_cases, "00B0F0")
create_sheet(wb, "16-举报处理", report_manage_cases, "FF0000")
create_sheet(wb, "17-公告管理", notice_manage_cases, "BF8F00")
create_sheet(wb, "18-文件上传", file_upload_cases, "92D050")

# 添加目录页
ws_index = wb.create_sheet(title="目录", index=0)
ws_index.cell(row=1, column=1, value="校园二手交易平台测试用例目录")
ws_index.cell(row=1, column=1).font = Font(bold=True, size=16)
ws_index.cell(row=1, column=1).alignment = Alignment(horizontal="center")

ws_index.cell(row=3, column=1, value="一、小程序用户端")
ws_index.cell(row=3, column=1).font = Font(bold=True, size=12, color="4472C4")

user_modules = [
    ("1-用户认证", "用户登录、注册、退出", len(user_auth_cases)),
    ("2-商品浏览", "商品列表、详情、搜索", len(goods_browse_cases)),
    ("3-商品发布", "发布、编辑、下架商品", len(goods_publish_cases)),
    ("4-商品购买", "下单、支付、取消订单", len(goods_purchase_cases)),
    ("5-收藏功能", "收藏、取消收藏", len(favorite_cases)),
    ("6-评价功能", "评论、查看评论", len(comment_cases)),
    ("7-举报功能", "举报商品", len(report_cases)),
    ("8-订单管理", "查看订单、订单详情", len(order_cases)),
    ("9-个人信息", "查看、编辑个人资料", len(profile_cases)),
]

row = 4
for sheet_name, desc, count in user_modules:
    ws_index.cell(row=row, column=1, value=sheet_name)
    ws_index.cell(row=row, column=2, value=desc)
    ws_index.cell(row=row, column=3, value=f"{count}条")
    row += 1

row += 1
ws_index.cell(row=row, column=1, value="二、管理后台")
ws_index.cell(row=row, column=1).font = Font(bold=True, size=12, color="C65911")
row += 1

admin_modules = [
    ("10-管理员登录", "管理员账号登录", len(admin_auth_cases)),
    ("11-数据统计", "仪表盘、趋势图", len(dashboard_cases)),
    ("12-商品审核", "审核通过/拒绝", len(goods_audit_cases)),
    ("13-用户管理", "查看、启用/禁用用户", len(user_manage_cases)),
    ("14-分类管理", "分类增删改查", len(category_manage_cases)),
    ("15-订单监管", "订单查看、监管", len(order_manage_cases)),
    ("16-举报处理", "举报审核、处理", len(report_manage_cases)),
    ("17-公告管理", "公告增删改查", len(notice_manage_cases)),
    ("18-文件上传", "文件上传功能", len(file_upload_cases)),
]

for sheet_name, desc, count in admin_modules:
    ws_index.cell(row=row, column=1, value=sheet_name)
    ws_index.cell(row=row, column=2, value=desc)
    ws_index.cell(row=row, column=3, value=f"{count}条")
    row += 1

ws_index.column_dimensions['A'].width = 20
ws_index.column_dimensions['B'].width = 25
ws_index.column_dimensions['C'].width = 10

total = sum([len(user_auth_cases), len(goods_browse_cases), len(goods_publish_cases),
             len(goods_purchase_cases), len(favorite_cases), len(comment_cases),
             len(report_cases), len(order_cases), len(profile_cases),
             len(admin_auth_cases), len(dashboard_cases), len(goods_audit_cases),
             len(user_manage_cases), len(category_manage_cases), len(order_manage_cases),
             len(report_manage_cases), len(notice_manage_cases), len(file_upload_cases)])

row += 2
ws_index.cell(row=row, column=1, value=f"测试用例总计：{total} 条")
ws_index.cell(row=row, column=1).font = Font(bold=True, size=14, color="FF0000")

output_path = "f:/Trae/project/second-hub-main/测试用例.xlsx"
wb.save(output_path)
print(f"测试用例已生成: {output_path}")
print(f"共计 {total} 条测试用例，分布在18个功能模块中")
